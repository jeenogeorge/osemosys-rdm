"""Aplica el par YES_PROP/DEP del Problema 2 (CapitalCost ↔ FixedCost agrupado).

Pre-requisito: Uncertainty_Table debe tener 15 filas con X_Num 1..15
(estado tras correr apply_rdm_cleanup.py).

Añade al final:
  X_Num=16  YES_PROP  CapitalCost   6 técnicas renovables agrupadas
  X_Num=17  DEP       FixedCost     mismas 6 técnicas (alineamiento idx-by-idx)

Idempotente: detecta si ya está aplicado y aborta sin tocar nada.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

XLSX = Path("src/Interface_RDM.xlsx")

RE_TECHS = "PWRSOL001 ; PWRWND001 ; PWRBIO001 ; PWRGEO ; PWRCSP001 ; PWRCSP002"


def col_index(ws, header_name: str) -> int:
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    raise KeyError(header_name)


def write_row(ws, row: int, data: dict) -> None:
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers):
        if h in data:
            ws.cell(row=row, column=i + 1, value=data[h])


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} no existe", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Uncertainty_Table"]
    xnum_col = col_index(ws, "X_Num")

    last_data_row = 1
    last_xnum = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=xnum_col).value
        if v is not None:
            last_data_row = r
            try:
                last_xnum = int(v)
            except (TypeError, ValueError):
                pass

    if last_xnum >= 17:
        print(f"Ya aplicado (último X_Num={last_xnum} >= 17). Aborto.")
        return 0
    if last_xnum != 15:
        print(f"ERROR: estado inesperado. Último X_Num={last_xnum}, esperado 15. "
              f"Corre apply_rdm_cleanup.py primero.", file=sys.stderr)
        return 2

    primary = {
        "X_Num": 16,
        "X_Category": "Capital Costs",
        "X_Plain_English_Description": "Renewable power capital costs",
        "Involved_Scenarios": "Scenario1",
        "X_Mathematical_Type": "Time_Series",
        "Explored_Parameter_of_X": "Final_Value_Multiplicative",
        "Initial_Year_of_Uncertainty": 2027,
        "Min_Value": 0.7,
        "Max_Value": 1.3,
        "Exact_Parameters_Involved_in_Osemosys": "CapitalCost",
        "Dependency": "YES_PROP",
        "Involved_First_Sets_in_Osemosys": RE_TECHS,
        "Involved_Second_Sets_in_Osemosys": "-",
        "Involved_Third_Sets_in_Osemosys": "-",
    }
    dependent = {
        "X_Num": 17,
        "X_Category": "Fixed Costs",
        "X_Plain_English_Description": "Renewable power fixed O&M costs (linked to capital)",
        "Involved_Scenarios": "Scenario1",
        "X_Mathematical_Type": "Time_Series",
        "Explored_Parameter_of_X": "Final_Value_Multiplicative",
        "Initial_Year_of_Uncertainty": 2027,
        "Min_Value": 0.7,
        "Max_Value": 1.3,
        "Exact_Parameters_Involved_in_Osemosys": "FixedCost",
        "Dependency": "DEP",
        "Involved_First_Sets_in_Osemosys": RE_TECHS,
        "Involved_Second_Sets_in_Osemosys": "-",
        "Involved_Third_Sets_in_Osemosys": "-",
    }

    write_row(ws, last_data_row + 1, primary)
    write_row(ws, last_data_row + 2, dependent)
    print(f"P2: añadidas filas X_Num=16 (YES_PROP CapitalCost) y X_Num=17 (DEP FixedCost) "
          f"con 6 técnicas renovables alineadas")

    wb.save(XLSX)
    print(f"OK: guardado {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
