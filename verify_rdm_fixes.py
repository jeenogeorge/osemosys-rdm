"""Verifica que las correcciones 4A-4D y Problema 2 se aplicaron correctamente."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

XLSX = Path("src/Interface_RDM.xlsx")
SCENARIO1 = Path("src/workflow/0_Scenarios/Scenario1.txt")


def col_index(ws, header_name: str) -> int:
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    raise KeyError(header_name)


def main() -> int:
    errors = []
    warnings = []

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Uncertainty_Table"]
    setup = wb["Setup"]

    xnum_col = col_index(ws, "X_Num")
    dep_col = col_index(ws, "Dependency")
    first_set_col = col_index(ws, "Involved_First_Sets_in_Osemosys")
    inv_scen_col = col_index(ws, "Involved_Scenarios")
    desc_col = col_index(ws, "X_Plain_English_Description")
    param_col = col_index(ws, "Exact_Parameters_Involved_in_Osemosys")

    rows = []
    for r in range(2, ws.max_row + 1):
        x = ws.cell(row=r, column=xnum_col).value
        if x is None:
            continue
        rows.append({
            "excel_row": r,
            "x_num": int(x),
            "dep": (ws.cell(row=r, column=dep_col).value or "").strip(),
            "first_set": str(ws.cell(row=r, column=first_set_col).value or ""),
            "scen": str(ws.cell(row=r, column=inv_scen_col).value or ""),
            "desc": str(ws.cell(row=r, column=desc_col).value or ""),
            "param": str(ws.cell(row=r, column=param_col).value or ""),
        })

    print(f"Total filas con datos: {len(rows)}")

    # Check 1: X_Num secuencial 1..N
    x_nums = [r["x_num"] for r in rows]
    expected = list(range(1, len(rows) + 1))
    if x_nums != expected:
        errors.append(f"X_Num no secuencial: {x_nums}")
    else:
        print(f"OK X_Num secuencial 1..{len(rows)}")

    # Check 2: huérfanos YES_PROP
    orphans = []
    for i, r in enumerate(rows):
        if r["dep"] == "YES_PROP":
            next_dep = rows[i + 1]["dep"] if i + 1 < len(rows) else "EOF"
            if next_dep != "DEP":
                orphans.append((r["x_num"], next_dep))
    if orphans:
        errors.append(f"YES_PROP huérfanos: {orphans}")
    else:
        print("OK sin YES_PROP huérfanos")

    # Check 3: alineamiento Involved_First_Sets en pares YES_*/DEP
    misaligned = []
    for i in range(len(rows) - 1):
        if rows[i]["dep"] in ("YES_PROP", "YES_ADD", "YES_ELAST") and rows[i + 1]["dep"] == "DEP":
            if rows[i]["first_set"] != rows[i + 1]["first_set"]:
                # Es ok que difieran salvo en el caso del Problema 2 (mismo grupo)
                misaligned.append((rows[i]["x_num"], rows[i + 1]["x_num"],
                                  rows[i]["first_set"][:60], rows[i + 1]["first_set"][:60]))
    if misaligned:
        warnings.append(f"Pares YES_*/DEP con first_set distinto: {misaligned}")
    else:
        print("OK todos los pares YES_*/DEP tienen first_set idéntico")

    # Check 4: Problema 2 — fila YES_PROP CapitalCost + DEP FixedCost con same tech list
    # Solo aplicable cuando las filas de P2 ya fueron añadidas (X_Num >= 17)
    p2_primary = next((r for r in rows if r["dep"] == "YES_PROP" and r["param"].strip() == "CapitalCost"), None)
    p2_dep = next((r for r in rows if r["dep"] == "DEP" and r["param"].strip() == "FixedCost"), None)
    if not p2_primary and not p2_dep:
        print("INFO Problema 2: filas no encontradas — apply_rdm_problem2.py aún no aplicado (esperado en cleanup-only)")
    elif not p2_primary or not p2_dep:
        errors.append(f"Problema 2: falta primary o dependent (primary={bool(p2_primary)}, dep={bool(p2_dep)})")
    elif p2_primary["first_set"] != p2_dep["first_set"]:
        errors.append(f"Problema 2: first_set distinto entre CapitalCost y FixedCost")
    else:
        print(f"OK Problema 2: X_Num {p2_primary['x_num']}↔{p2_dep['x_num']} alineados ({p2_primary['first_set'][:60]}...)")

    # Check 5: 4C — no quedan referencias a Scenario2
    scen2_rows = [r["x_num"] for r in rows if "Scenario2" in r["scen"]]
    if scen2_rows:
        errors.append(f"4C: quedan filas con Scenario2: {scen2_rows}")
    else:
        print("OK sin referencias a Scenario2 en Involved_Scenarios")

    # Check 6: 4D — descripciones filas 1-4
    expected_desc = {
        1: "Demand Growth - livestock cattle",
        2: "Demand Growth - livestock other",
        3: "Demand Growth - livestock goat",
        4: "Demand Growth - livestock sheep",
    }
    for x, want in expected_desc.items():
        got = next((r["desc"] for r in rows if r["x_num"] == x), None)
        if got != want:
            errors.append(f"4D: X_Num={x} desc={got!r} esperado {want!r}")
    if not [e for e in errors if e.startswith("4D")]:
        print("OK descripciones filas 1-4 corregidas")

    # Check 7: 4B — Setup.Region == set REGION de Scenario1.txt
    region_col = col_index(setup, "Region")
    setup_region = str(setup.cell(row=2, column=region_col).value).strip()
    with open(SCENARIO1) as fh:
        txt = fh.read()
    m = re.search(r"set\s+REGION\s*:=\s*([^;]+);", txt)
    if not m:
        errors.append("No se encontró 'set REGION' en Scenario1.txt")
    else:
        scen_region = m.group(1).strip()
        if setup_region != scen_region:
            errors.append(f"4B: Setup.Region={setup_region!r} != Scenario1.set REGION={scen_region!r}")
        else:
            print(f"OK Setup.Region={setup_region!r} alineado con Scenario1.txt")

    # Resumen
    print()
    if errors:
        print(f"FALLOS ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")
    if warnings:
        print(f"WARNINGS ({len(warnings)}):")
        for w in warnings:
            print(f"  - {w}")
    if not errors:
        print("TODAS LAS VERIFICACIONES PASAN")
        return 0
    return 1


if __name__ == "__main__":
    sys.exit(main())
