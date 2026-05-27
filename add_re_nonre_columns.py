"""Añade tres columnas opt-in para acoplamiento RE/non-RE a Uncertainty_Table.

  RE_Techs       lista de técnicas renovables (separadas por ';')
  NonRE_Techs    lista de técnicas no-renovables (mismo formato)
  Sum_To_Value   valor target de la suma (por defecto 1.0 cuando RE/NonRE
                 están populados pero esta celda está vacía)

Cuando una fila tiene RE_Techs Y NonRE_Techs populados, el manager invoca
post-perturbación un fixer que reescribe cada NonRE_Tech con el valor
uniforme (Sum_To_Value - sum(RE_values)) / len(NonRE_Techs) por (R, Y).

Idempotente: si las columnas ya existen, no hace nada.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

XLSX = Path("src/Interface_RDM.xlsx")
NEW_COLS = ["RE_Techs", "NonRE_Techs", "Sum_To_Value"]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Uncertainty_Table"]

    existing = [c.value for c in ws[1]]
    next_col = ws.max_column + 1
    added = []
    for name in NEW_COLS:
        if name in existing:
            print(f"OK: columna {name!r} ya existe, skip")
            continue
        ws.cell(row=1, column=next_col, value=name)
        added.append((name, next_col))
        next_col += 1

    if added:
        wb.save(XLSX)
        for name, col in added:
            print(f"Añadida columna {name!r} en columna Excel {col} (vacía)")
    else:
        print("Nada que hacer")
    return 0


if __name__ == "__main__":
    sys.exit(main())
