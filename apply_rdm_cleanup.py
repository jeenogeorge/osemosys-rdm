"""Aplica fixes 4A, 4B, 4C, 4D a src/Interface_RDM.xlsx.

  4B  Setup.Region: BWA -> RE1
  4C  Uncertainty_Table.Involved_Scenarios: "Scenario1 ; Scenario2" -> "Scenario1"
  4D  Uncertainty_Table.X_Plain_English_Description filas 2-4: descripciones correctas
  4A  Insertar filas DEP para LVSGOACU y LVSSHPCU; renumerar X_Num 1..15

Idempotente: detecta si ya está aplicado y aborta sin tocar nada.
"""
from __future__ import annotations

import copy
import sys
from pathlib import Path

import openpyxl

XLSX = Path("src/Interface_RDM.xlsx")
BACKUP = Path("src/Interface_RDM.xlsx.bak")

DESCRIPTIONS = {
    1: "Demand Growth - livestock cattle",
    2: "Demand Growth - livestock other",
    3: "Demand Growth - livestock goat",
    4: "Demand Growth - livestock sheep",
}


def col_index(ws, header_name: str) -> int:
    for cell in ws[1]:
        if cell.value == header_name:
            return cell.column
    raise KeyError(f"Column {header_name!r} not found in sheet {ws.title!r}")


def read_row(ws, row: int) -> dict:
    headers = [c.value for c in ws[1]]
    return {h: ws.cell(row=row, column=i + 1).value for i, h in enumerate(headers)}


def write_row(ws, row: int, data: dict) -> None:
    headers = [c.value for c in ws[1]]
    for i, h in enumerate(headers):
        if h in data:
            ws.cell(row=row, column=i + 1, value=data[h])


def find_row_by_xnum(ws, x_num) -> int | None:
    col = col_index(ws, "X_Num")
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None:
            continue
        try:
            if int(v) == int(x_num):
                return r
        except (TypeError, ValueError):
            continue
    return None


def main() -> int:
    if not XLSX.exists():
        print(f"ERROR: {XLSX} no existe", file=sys.stderr)
        return 1
    if not BACKUP.exists():
        print(f"ERROR: backup {BACKUP} no existe — crea uno antes de correr", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(XLSX)
    ws_unc = wb["Uncertainty_Table"]
    ws_setup = wb["Setup"]

    setup_region = ws_setup.cell(row=2, column=col_index(ws_setup, "Region")).value
    last_xnum = None
    xnum_col = col_index(ws_unc, "X_Num")
    for r in range(2, ws_unc.max_row + 1):
        v = ws_unc.cell(row=r, column=xnum_col).value
        if v is not None:
            try:
                last_xnum = int(v)
            except (TypeError, ValueError):
                pass

    print(f"Estado inicial: Setup.Region={setup_region!r}, último X_Num={last_xnum}")

    if setup_region == "RE1" and last_xnum is not None and last_xnum >= 15:
        print("Ya aplicado (Region=RE1 y X_Num>=15). Aborto sin cambios.")
        return 0

    # ---------- 4B: Setup.Region BWA -> RE1 ----------
    region_col = col_index(ws_setup, "Region")
    old_region = ws_setup.cell(row=2, column=region_col).value
    ws_setup.cell(row=2, column=region_col, value="RE1")
    print(f"4B: Setup.Region {old_region!r} -> 'RE1'")

    # ---------- 4C: limpiar "Scenario1 ; Scenario2" -> "Scenario1" ----------
    inv_scen_col = col_index(ws_unc, "Involved_Scenarios")
    n_4c = 0
    for r in range(2, ws_unc.max_row + 1):
        v = ws_unc.cell(row=r, column=inv_scen_col).value
        if v is None:
            continue
        if "Scenario2" in str(v):
            ws_unc.cell(row=r, column=inv_scen_col, value="Scenario1")
            n_4c += 1
    print(f"4C: limpiadas {n_4c} filas de Involved_Scenarios")

    # ---------- 4D: descripciones correctas filas 2-4 ----------
    desc_col = col_index(ws_unc, "X_Plain_English_Description")
    for x_num, desc in DESCRIPTIONS.items():
        r = find_row_by_xnum(ws_unc, x_num)
        if r is None:
            print(f"WARN: 4D no encontró X_Num={x_num}", file=sys.stderr)
            continue
        old = ws_unc.cell(row=r, column=desc_col).value
        if old != desc:
            ws_unc.cell(row=r, column=desc_col, value=desc)
            print(f"4D: X_Num={x_num} desc {old!r} -> {desc!r}")

    # ---------- 4A: insertar filas DEP, renumerar 1..15 ----------
    row_goa_cf = find_row_by_xnum(ws_unc, 12)
    row_shp_cf = find_row_by_xnum(ws_unc, 13)
    if row_goa_cf is None or row_shp_cf is None:
        print(f"ERROR: 4A no encontró filas template (12/13)", file=sys.stderr)
        return 2

    template_goa = read_row(ws_unc, row_goa_cf)
    template_shp = read_row(ws_unc, row_shp_cf)

    dep_goa = copy.deepcopy(template_goa)
    dep_goa["X_Plain_English_Description"] = "WATER FOR GOAT - TYPE CU"
    dep_goa["Involved_First_Sets_in_Osemosys"] = "LVSGOACU"
    dep_goa["Dependency"] = "DEP"

    dep_shp = copy.deepcopy(template_shp)
    dep_shp["X_Plain_English_Description"] = "WATER FOR SHEEP - TYPE CU"
    dep_shp["Involved_First_Sets_in_Osemosys"] = "LVSSHPCU"
    dep_shp["Dependency"] = "DEP"

    ws_unc.insert_rows(row_shp_cf + 1)
    write_row(ws_unc, row_shp_cf + 1, dep_shp)
    print(f"4A: insertada fila DEP LVSSHPCU después de la fila {row_shp_cf}")

    ws_unc.insert_rows(row_goa_cf + 1)
    write_row(ws_unc, row_goa_cf + 1, dep_goa)
    print(f"4A: insertada fila DEP LVSGOACU después de la fila {row_goa_cf}")

    n_data = 0
    for r in range(2, ws_unc.max_row + 1):
        x_num = ws_unc.cell(row=r, column=xnum_col).value
        if x_num is None:
            continue
        n_data += 1
        ws_unc.cell(row=r, column=xnum_col, value=n_data)
    print(f"4A: renumerado X_Num 1..{n_data} ({n_data} filas con datos)")

    wb.save(XLSX)
    print(f"OK: guardado {XLSX}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
