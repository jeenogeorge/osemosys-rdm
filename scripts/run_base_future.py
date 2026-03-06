# -*- coding: utf-8 -*-
"""
Base Future Execution Wrapper for DVC Pipeline

This script executes only the base future (Future 0) scenario by temporarily
modifying the Interface_RDM.xlsx configuration to run only the base case.

Usage:
    python scripts/run_base_future.py

Manual execution:
    This script can be run independently for debugging or testing base futures.

DVC integration:
    Called automatically by DVC when executing the 'base_future' stage.

Author: AFR_RDM Team
"""

import os
import sys
import json
import time
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Add src to path to import RUN_RDM
sys.path.insert(0, str(Path(__file__).parent.parent / 'src'))

def backup_interface(interface_path):
    """Create a backup of Interface_RDM.xlsx"""
    backup_path = interface_path.with_suffix('.xlsx.backup')
    shutil.copy2(interface_path, backup_path)
    return backup_path

def restore_interface(interface_path, backup_path):
    """Restore Interface_RDM.xlsx from backup"""
    if backup_path.exists():
        shutil.copy2(backup_path, interface_path)
        backup_path.unlink()

def configure_for_base_only(interface_path):
    """
    Modify Interface_RDM.xlsx to run only base future.
    Sets Run_Base_Future=Yes and Run_RDM=No
    Also caches formula values in all sheets to avoid NaN issues when
    openpyxl saves the file (openpyxl cannot evaluate Excel formulas).
    """
    print("📝 Configuring Interface_RDM.xlsx for base future only...")

    # Load with data_only=True to get calculated values from formulas
    wb_data = load_workbook(interface_path, data_only=True)

    # Extract calculated values from ALL sheets (formulas exist in
    # Uncertainty_Table, Params_Sets_Vari, and potentially others)
    all_cached_values = {}  # {sheet_name: {(row, col): value}}
    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        sheet_cache = {}
        for row_idx in range(1, ws_data.max_row + 1):
            for col_idx in range(1, ws_data.max_column + 1):
                cell_value = ws_data.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    sheet_cache[(row_idx, col_idx)] = cell_value
        if sheet_cache:
            all_cached_values[sheet_name] = sheet_cache
    wb_data.close()

    # Load normally to modify
    wb = load_workbook(interface_path)

    # Modify Setup sheet
    ws_setup = wb['Setup']
    headers = {}
    for col_idx, cell in enumerate(ws_setup[1], start=1):
        if cell.value:
            headers[cell.value] = col_idx

    if 'Run_Base_Future' in headers:
        ws_setup.cell(row=2, column=headers['Run_Base_Future'], value='Yes')
    if 'Run_RDM' in headers:
        ws_setup.cell(row=2, column=headers['Run_RDM'], value='No')

    # Write cached values back to ALL sheets to preserve formula results
    for sheet_name, sheet_cache in all_cached_values.items():
        ws = wb[sheet_name]
        for (row_idx, col_idx), value in sheet_cache.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            # Only replace formula cells with their calculated values
            if cell.value and str(cell.value).startswith('='):
                cell.value = value

    # Save the workbook
    wb.save(interface_path)
    wb.close()

    print("✓ Configuration updated: Run_Base_Future=Yes, Run_RDM=No")

def generate_metrics(executables_dir, metrics_file):
    """
    Generate metrics JSON file for DVC tracking.
    Extracts execution time and basic statistics.
    """
    metrics = {
        "stage": "base_future",
        "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
        "scenarios_processed": 0,
        "total_files": 0
    }

    # Count processed scenarios
    if executables_dir.exists():
        scenario_dirs = [d for d in executables_dir.iterdir() if d.is_dir() and '_0' in d.name]
        metrics["scenarios_processed"] = len(scenario_dirs)

        # Count output files
        for scenario_dir in scenario_dirs:
            csv_files = list(scenario_dir.glob('*.csv'))
            metrics["total_files"] += len(csv_files)

    # Write metrics
    metrics_file.parent.mkdir(parents=True, exist_ok=True)
    with open(metrics_file, 'w') as f:
        json.dump(metrics, f, indent=2)

    print(f"✓ Metrics written to {metrics_file}")
    print(f"  - Scenarios processed: {metrics['scenarios_processed']}")
    print(f"  - Total output files: {metrics['total_files']}")

def main():
    """Main execution function"""
    print("=" * 70)
    print("AFR_RDM - Base Future Execution")
    print("=" * 70)

    # Paths
    project_root = Path(__file__).parent.parent
    interface_path = project_root / 'src' / 'Interface_RDM.xlsx'
    executables_dir = project_root / 'src' / 'workflow' / '1_Experiment' / 'Executables'
    metrics_file = project_root / 'src' / 'workflow' / '1_Experiment' / 'base_future_metrics.json'

    # Verify Interface_RDM.xlsx exists
    if not interface_path.exists():
        print(f"❌ Error: {interface_path} not found")
        sys.exit(1)

    # Backup and configure
    backup_path = None
    start_time = time.time()

    try:
        # Backup original configuration
        backup_path = backup_interface(interface_path)
        print(f"✓ Backup created: {backup_path.name}")

        # Configure for base future only
        configure_for_base_only(interface_path)

        # Change to src directory and execute RUN_RDM
        print("\n🚀 Executing RUN_RDM.py...")
        print("-" * 70)

        os.chdir(project_root / 'src')

        # Import and execute (this preserves the original RUN_RDM.py logic)
        import RUN_RDM

        print("-" * 70)

        # Generate metrics
        elapsed_time = time.time() - start_time
        print(f"\n⏱️  Execution time: {elapsed_time:.2f} seconds")

        generate_metrics(executables_dir, metrics_file)

        print("\n✅ Base future execution completed successfully!")

    except Exception as e:
        print(f"\n❌ Error during execution: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    finally:
        # Always restore original configuration
        if backup_path:
            os.chdir(project_root)
            restore_interface(interface_path, backup_path)
            print(f"✓ Configuration restored from backup")

if __name__ == "__main__":
    main()
