"""Añade columnas RE_Param y NonRE_Param a la hoja Setup de Interface_RDM.xlsx.

Las deja vacías: el fixer de renewable share NO se activa hasta que el usuario
populate ambas columnas con nombres de parámetros válidos del modelo.
Idempotente: si las columnas ya existen, no hace nada.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

XLSX = Path("src/Interface_RDM.xlsx")
NEW_COLS = ["RE_Param", "NonRE_Param"]


def main() -> int:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Setup"]

    existing = [c.value for c in ws[1]]
    next_col = ws.max_column + 1
    added = []
    for name in NEW_COLS:
        if name in existing:
            print(f"OK: columna {name!r} ya existe, skip")
            continue
        ws.cell(row=1, column=next_col, value=name)
        # Dejar la fila 2 vacía (None) — comportamiento: fixer inactivo
        added.append((name, next_col))
        next_col += 1

    if added:
        wb.save(XLSX)
        for name, col in added:
            print(f"Añadida columna {name!r} en columna Excel {col} (vacía)")
    else:
        print("Nada que hacer")
    return 0


if __name__ == "__main__":
    sys.exit(main())
